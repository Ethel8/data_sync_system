import io
import openpyxl
from flask import Blueprint, request, Response
from app import db
from models import Order

export_bp = Blueprint('export', __name__)


@export_bp.route('/orders', methods=['GET'])
def export_orders():
    """导出订单数据为 Excel"""
    order_status = request.args.get('status', '', type=str)
    order_progress = request.args.get('progress', '', type=str)
    keyword = request.args.get('keyword', '', type=str)

    query = Order.query
    if order_status:
        query = query.filter_by(order_status=order_status)
    if order_progress:
        query = query.filter_by(order_progress=order_progress)
    if keyword:
        query = query.filter(
            db.or_(
                Order.sales_order_no.contains(keyword),
                Order.ship_to_name.contains(keyword),
                Order.fara_external_code.contains(keyword),
            )
        )

    orders = query.order_by(Order.updated_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '订单数据'

    headers = [
        '销售订单号', '送达方名称', '创建日期', '法拉外码', '订单数量',
        '发货日期', '已发货数量', '公司开票数量', '客户开票数量',
        '买入单价不含税', '客户应付含税金额', '负责人',
        '订单状态', '订单进度', '公司开票', '客户开票',
    ]
    ws.append(headers)

    # 表头样式
    from openpyxl.styles import Font, PatternFill
    bold = Font(bold=True)
    fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    white_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.font = white_font
        cell.fill = fill

    for order in orders:
        ws.append([
            order.sales_order_no,
            order.ship_to_name,
            order.create_date or '',
            order.fara_external_code,
            order.order_quantity or 0,
            order.ship_date or '',
            order.shipped_quantity or 0,
            order.factory_invoice_quantity or 0,
            order.customer_invoice_quantity or 0,
            order.purchase_price_excl_tax or 0,
            order.customer_payable_incl_tax or 0,
            order.manager or '',
            order.order_status or '',
            order.order_progress or '',
            order.factory_invoice_status or '',
            order.customer_invoice_status or '',
        ])

    # 自动列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=orders_export.xlsx'}
    )
